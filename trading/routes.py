"""
回测API路由 - 提供RESTful接口
"""
from flask import Blueprint, request, jsonify, send_file
from trading.backtest_dao import BacktestDAO
from trading.backtest_engine import BacktestEngine
from trading.backtest_batch_queue import BacktestBatchQueue
from utils.db_manager import DBManager
from utils.akshare_fetcher import AKShareFetcher
from utils.strategy_name_mapper import get_english_name
from utils.strategy_config_manager import StrategyConfigManager
import logging

# 获取日志记录器
logger = logging.getLogger(__name__)

# 创建蓝图
trading_bp = Blueprint('trading', __name__, url_prefix='/api/trading')

# ==================== 回测相关接口 ====================

# 初始化回测相关组件
backtest_dao = BacktestDAO()
from utils.global_db import get_global_db
db_manager = get_global_db()
akshare_fetcher = AKShareFetcher("data")


@trading_bp.route('/backtest/configs', methods=['GET'])
def get_backtest_configs():
    """
    获取回测配置列表接口
    
    返回:
        {
            "success": true/false,
            "message": "成功或错误信息",
            "data": {
                "configs": [
                    {
                        "id": 1,
                        "config_name": "测试配置",
                        "strategy_name": "多方炮策略",
                        "score_threshold": 60,
                        "hold_period": 10,
                        "stop_loss": -5,
                        "take_profit": 15,
                        "initial_capital": 300000,
                        "buy_amount": 100000,
                        "max_daily_buys": 8,
                        "support_level_method": "ma20",
                        "buy_point_lower": -1,
                        "buy_point_upper": 3,
                        "start_date": "2024-01-01",
                        "end_date": "2024-06-30",
                        "created_at": "2024-01-01 12:00:00"
                    }
                ],
                "total_count": 1
            }
        }
    """
    try:
        # 调用DAO获取所有配置
        configs = backtest_dao.get_all_configs()
        
        return jsonify({
            'success': True,
            'message': '获取回测配置列表成功',
            'data': {
                'configs': configs,
                'total_count': len(configs)
            }
        }), 200
    
    except Exception as e:
        logger.error(f"获取回测配置列表失败: {str(e)}")
        return jsonify({
            'success': False,
            'message': f'获取回测配置列表失败: {str(e)}',
            'data': None
        }), 500


# ==================== 批量回测任务队列接口 ====================


@trading_bp.route('/backtest/batch/submit', methods=['POST'])
def submit_batch_backtest():
    """
    提交批量回测任务

    请求体:
        {
            "tasks": [
                {
                    "strategy_name": "涨停横盘策略",
                    "start_date": "2026-01-01",
                    "end_date": "2026-03-31",
                    "timing_strategy": "turtle",
                    "support_level_method": "ma20"
                }
            ],
            "config": {
                "initial_capital": 300000,
                "score_threshold": 60,
                "max_daily_buys": 3
            }
        }

    返回:
        {
            "success": true/false,
            "data": {
                "batch_id": "batch_xxx",
                "total_tasks": 5
            }
        }
    """
    try:
        data = request.get_json() or {}

        tasks = data.get('tasks', [])
        config = data.get('config', {})

        if not tasks:
            return jsonify({
                'success': False,
                'message': '任务列表为空',
                'data': None
            }), 400

        # 创建批量任务队列
        batch_queue = BacktestBatchQueue.create_batch(tasks, config)

        logger.info(f"提交批量回测任务: {batch_queue.batch_id}, 任务数: {len(tasks)}")

        return jsonify({
            'success': True,
            'message': '批量任务提交成功',
            'data': {
                'batch_id': batch_queue.batch_id,
                'total_tasks': len(tasks)
            }
        })

    except Exception as e:
        logger.error(f"提交批量回测任务失败: {str(e)}")
        return jsonify({
            'success': False,
            'message': f'提交批量回测任务失败: {str(e)}',
            'data': None
        }), 500


@trading_bp.route('/backtest/batch/start', methods=['POST'])
def start_batch_backtest():
    """
    开始批量回测执行

    请求体:
        {
            "batch_id": "batch_xxx"
        }

    返回:
        {
            "success": true/false,
            "message": "批量任务已开始执行"
        }
    """
    try:
        data = request.get_json() or {}
        batch_id = data.get('batch_id')

        if not batch_id:
            return jsonify({
                'success': False,
                'message': '缺少 batch_id',
                'data': None
            }), 400

        # 获取执行器
        batch_queue = BacktestBatchQueue.get_executor(batch_id)

        if not batch_queue:
            return jsonify({
                'success': False,
                'message': '批量任务不存在',
                'data': None
            }), 404

        if batch_queue.status == 'running':
            return jsonify({
                'success': False,
                'message': '批量任务已在执行中',
                'data': None
            }), 400

        # 启动执行
        batch_queue.start()

        logger.info(f"开始批量回测执行: {batch_id}")

        return jsonify({
            'success': True,
            'message': '批量任务已开始执行',
            'data': {
                'batch_id': batch_id
            }
        })

    except Exception as e:
        logger.error(f"开始批量回测执行失败: {str(e)}")
        return jsonify({
            'success': False,
            'message': f'开始批量回测执行失败: {str(e)}',
            'data': None
        }), 500


@trading_bp.route('/backtest/batch/status', methods=['GET'])
def get_batch_backtest_status():
    """
    查询批量回测执行状态

    参数:
        batch_id: 批量任务ID (query参数)

    返回:
        {
            "success": true/false,
            "data": {
                "batch_id": "batch_xxx",
                "status": "running",
                "total_tasks": 5,
                "completed_tasks": 2,
                "failed_tasks": 0,
                "current_task": {
                    "index": 2,
                    "strategy_name": "涨停横盘策略",
                    "status": "running"
                }
            }
        }
    """
    try:
        batch_id = request.args.get('batch_id')

        if not batch_id:
            return jsonify({
                'success': False,
                'message': '缺少 batch_id',
                'data': None
            }), 400

        # 获取执行器
        batch_queue = BacktestBatchQueue.get_executor(batch_id)

        if not batch_queue:
            return jsonify({
                'success': False,
                'message': '批量任务不存在',
                'data': None
            }), 404

        # 获取状态
        status = batch_queue.get_status()

        return jsonify({
            'success': True,
            'data': status
        })

    except Exception as e:
        logger.error(f"查询批量回测状态失败: {str(e)}")
        return jsonify({
            'success': False,
            'message': f'查询批量回测状态失败: {str(e)}',
            'data': None
        }), 500


@trading_bp.route('/backtest/batch/results', methods=['GET'])
def get_batch_backtest_results():
    """
    获取批量回测执行结果

    参数:
        batch_id: 批量任务ID (query参数)

    返回:
        {
            "success": true/false,
            "data": {
                "batch_id": "batch_xxx",
                "status": "completed",
                "results": [...]
            }
        }
    """
    try:
        batch_id = request.args.get('batch_id')

        if not batch_id:
            return jsonify({
                'success': False,
                'message': '缺少 batch_id',
                'data': None
            }), 400

        # 获取执行器
        batch_queue = BacktestBatchQueue.get_executor(batch_id)

        if not batch_queue:
            return jsonify({
                'success': False,
                'message': '批量任务不存在',
                'data': None
            }), 404

        # 获取结果
        results = batch_queue.get_results()

        return jsonify({
            'success': True,
            'data': results
        })

    except Exception as e:
        logger.error(f"获取批量回测结果失败: {str(e)}")
        return jsonify({
            'success': False,
            'message': f'获取批量回测结果失败: {str(e)}',
            'data': None
        }), 500


@trading_bp.route('/backtest/batch/cancel', methods=['POST'])
def cancel_batch_backtest():
    """
    取消批量回测执行

    请求体:
        {
            "batch_id": "batch_xxx"
        }

    返回:
        {
            "success": true/false,
            "message": "批量任务已取消"
        }
    """
    try:
        data = request.get_json() or {}
        batch_id = data.get('batch_id')

        if not batch_id:
            return jsonify({
                'success': False,
                'message': '缺少 batch_id',
                'data': None
            }), 400

        # 获取执行器
        batch_queue = BacktestBatchQueue.get_executor(batch_id)

        if not batch_queue:
            return jsonify({
                'success': False,
                'message': '批量任务不存在',
                'data': None
            }), 404

        # 停止执行
        batch_queue.stop()

        logger.info(f"取消批量回测执行: {batch_id}")

        return jsonify({
            'success': True,
            'message': '批量任务已取消',
            'data': {
                'batch_id': batch_id
            }
        })

    except Exception as e:
        logger.error(f"取消批量回测执行失败: {str(e)}")
        return jsonify({
            'success': False,
            'message': f'取消批量回测执行失败: {str(e)}',
            'data': None
        }), 500


@trading_bp.route('/backtest/batch/list', methods=['GET'])
def list_batch_backtest():
    """
    获取批量回测任务列表

    返回:
        {
            "success": true/false,
            "data": {
                "batches": [...]
            }
        }
    """
    try:
        from pathlib import Path

        batch_queue_dir = Path("data/backtest_batch")
        batches = []

        if batch_queue_dir.exists():
            for queue_file in batch_queue_dir.glob("queue_*.json"):
                try:
                    with open(queue_file, 'r', encoding='utf-8') as f:
                        data = json.load(f)
                        batches.append({
                            'batch_id': data.get('batch_id'),
                            'status': data.get('status'),
                            'total_tasks': len(data.get('tasks', [])),
                            'created_at': data.get('created_at'),
                            'started_at': data.get('started_at'),
                            'completed_at': data.get('completed_at')
                        })
                except Exception as e:
                    logger.warning(f"读取批量任务文件失败: {queue_file}, error: {str(e)}")

        # 按创建时间倒序排列
        batches.sort(key=lambda x: x.get('created_at', ''), reverse=True)

        return jsonify({
            'success': True,
            'data': {
                'batches': batches
            }
        })

    except Exception as e:
        logger.error(f"获取批量回测任务列表失败: {str(e)}")
        return jsonify({
            'success': False,
            'message': f'获取批量回测任务列表失败: {str(e)}',
            'data': None
        }), 500


@trading_bp.route('/backtest/configs/<int:config_id>', methods=['GET'])
def get_backtest_config(config_id):
    """
    获取单个回测配置详情接口
    
    参数:
        config_id: 配置ID (路径参数)
    
    返回:
        {
            "success": true/false,
            "message": "成功或错误信息",
            "data": {
                "id": 1,
                "config_name": "测试配置",
                "strategy_name": "多方炮策略",
                "score_threshold": 60,
                "hold_period": 10,
                "stop_loss": -5,
                "take_profit": 15,
                "initial_capital": 300000,
                "buy_amount": 100000,
                "max_daily_buys": 8,
                "support_level_method": "ma20",
                "buy_point_lower": -1,
                "buy_point_upper": 3,
                "start_date": "2024-01-01",
                "end_date": "2024-06-30",
                "created_at": "2024-01-01 12:00:00"
            }
        }
    """
    try:
        # 获取回测配置
        config = backtest_dao.get_config(config_id)
        if not config:
            return jsonify({
                'success': False,
                'message': '配置不存在',
                'data': None
            }), 404
        
        return jsonify({
            'success': True,
            'message': '获取回测配置详情成功',
            'data': config
        }), 200
    
    except Exception as e:
        logger.error(f"获取回测配置详情失败: {str(e)}")
        return jsonify({
            'success': False,
            'message': f'获取回测配置详情失败: {str(e)}',
            'data': None
        }), 500


@trading_bp.route('/backtest/configs', methods=['POST'])
def create_backtest_config():
    """
    创建回测配置接口
    
    请求体:
        {
            "config_name": "测试配置",
            "strategy_name": "多方炮策略",
            "score_threshold": 60,
            "hold_period": 10,
            "stop_loss": -5,
            "take_profit": 15,
            "initial_capital": 300000,
            "buy_amount": 100000,
            "max_daily_buys": 8,
            "support_level_method": "ma20",
            "buy_point_lower": -1,
            "buy_point_upper": 3,
            "start_date": "2024-01-01",
            "end_date": "2024-06-30"
        }
    
    返回:
        {
            "success": true/false,
            "message": "成功或错误信息",
            "data": {
                "config_id": 1
            }
        }
    """
    try:
        # 获取请求数据
        data = request.get_json() or {}
        
        # 验证必需参数（只验证config_name）
        if 'config_name' not in data or data['config_name'] is None:
            return jsonify({
                'success': False,
                'message': '缺少必需参数: config_name',
                'data': None
            }), 400
        
        # 调用DAO保存配置
        config_id = backtest_dao.save_config(data)
        
        if config_id == 0:
            return jsonify({
                'success': False,
                'message': '创建回测配置失败',
                'data': None
            }), 400
        
        return jsonify({
            'success': True,
            'message': '创建回测配置成功',
            'data': {
                'config_id': config_id
            }
        }), 200
    
    except Exception as e:
        logger.error(f"创建回测配置失败: {str(e)}")
        return jsonify({
            'success': False,
            'message': f'创建回测配置失败: {str(e)}',
            'data': None
        }), 500


@trading_bp.route('/backtest/configs/<int:config_id>', methods=['PUT'])
def update_backtest_config(config_id):
    """
    更新回测配置接口
    
    参数:
        config_id: 配置ID (路径参数)
    
    请求体:
        {
            "config_name": "测试配置",
            "strategy_name": "多方炮策略",
            "score_threshold": 60,
            "hold_period": 10,
            "stop_loss": -5,
            "take_profit": 15,
            "initial_capital": 300000,
            "buy_amount": 100000,
            "max_daily_buys": 8,
            "support_level_method": "ma20",
            "buy_point_lower": -1,
            "buy_point_upper": 3,
            "start_date": "2024-01-01",
            "end_date": "2024-06-30"
        }
    
    返回:
        {
            "success": true/false,
            "message": "成功或错误信息",
            "data": {
                "config_id": 1
            }
        }
    """
    try:
        # 获取请求数据
        data = request.get_json() or {}
        
        # 调用DAO更新配置
        success = backtest_dao.update_config(config_id, data)
        
        if not success:
            return jsonify({
                'success': False,
                'message': '更新回测配置失败',
                'data': None
            }), 400
        
        return jsonify({
            'success': True,
            'message': '更新回测配置成功',
            'data': {
                'config_id': config_id
            }
        }), 200
    
    except Exception as e:
        logger.error(f"更新回测配置失败: {str(e)}")
        return jsonify({
            'success': False,
            'message': f'更新回测配置失败: {str(e)}',
            'data': None
        }), 500


@trading_bp.route('/backtest/configs/<int:config_id>', methods=['DELETE'])
def delete_backtest_config(config_id):
    """
    删除回测配置接口
    
    参数:
        config_id: 配置ID (路径参数)
    
    返回:
        {
            "success": true/false,
            "message": "成功或错误信息",
            "data": {
                "config_id": 1
            }
        }
    """
    try:
        # 调用DAO删除配置
        success = backtest_dao.delete_config(config_id)
        
        if not success:
            return jsonify({
                'success': False,
                'message': '删除回测配置失败',
                'data': None
            }), 400
        
        return jsonify({
            'success': True,
            'message': '删除回测配置成功',
            'data': {
                'config_id': config_id
            }
        }), 200
    
    except Exception as e:
        logger.error(f"删除回测配置失败: {str(e)}")
        return jsonify({
            'success': False,
            'message': f'删除回测配置失败: {str(e)}',
            'data': None
        }), 500


@trading_bp.route('/backtest/results/<int:result_id>', methods=['DELETE'])
def delete_backtest_result(result_id):
    """
    删除回测结果接口
    
    参数:
        result_id: 回测结果ID (路径参数)
    
    返回:
        {
            "success": true/false,
            "message": "成功或错误信息",
            "data": {
                "result_id": 1
            }
        }
    """
    try:
        # 调用DAO删除回测结果（级联删除关联的收益曲线和交易记录）
        success = backtest_dao.delete_result(result_id)
        
        if not success:
            return jsonify({
                'success': False,
                'message': '删除回测结果失败',
                'data': None
            }), 400
        
        return jsonify({
            'success': True,
            'message': '删除回测结果成功',
            'data': {
                'result_id': result_id
            }
        }), 200
    
    except Exception as e:
        logger.error(f"删除回测结果失败: {str(e)}")
        return jsonify({
            'success': False,
            'message': f'删除回测结果失败: {str(e)}',
            'data': None
        }), 500


@trading_bp.route('/backtest/run', methods=['POST'])
def run_backtest():
    """
    运行回测接口
    
    请求体:
        {
            "strategy_name": "连阳回调策略",  # 中文名称
            "support_level_method": "ma20",
            "start_date": "2024-01-01",
            "end_date": "2024-06-30"
        }
    
    返回:
        {
            "success": true/false,
            "message": "成功或错误信息",
            "data": {
                "result_id": 1,
                "strategy_name": "连阳回调策略",  # 中文名称
                "total_return": 10.5,
                "win_rate": 65.0,
                "max_drawdown": 8.2,
                "sharpe_ratio": 1.2
            }
        }
    """
    try:
        # 获取请求数据
        data = request.get_json() or {}
        
        # 提取执行条件和回测配置
        strategy_name = data.get('strategy_name', '')  # 接收中文名称
        support_level_method = data.get('support_level_method', 'ma20')
        timing_strategy = data.get('timing_strategy', 'turtle')
        timing_params = data.get('timing_params', {})  # 提取择时策略参数
        start_date = data.get('start_date', '')
        end_date = data.get('end_date', '')
        
        # 将中文策略名称转换为英文（用于策略执行）
        english_strategy_name = get_english_name(strategy_name)
        
        # 从数据库读取回测配置参数
        db_config = db_manager.query_one("SELECT stop_loss, take_profit, hold_period, initial_capital, buy_amount, max_daily_buys, score_threshold FROM backtest_config LIMIT 1")
        
        # 使用数据库中的配置值
        if db_config:
            score_threshold = db_config.get('score_threshold', 60)
            max_hold_days = db_config.get('hold_period', 10)
            stop_loss = db_config.get('stop_loss', -7)
            take_profit = db_config.get('take_profit', 21)
            initial_capital = db_config.get('initial_capital', 300000)
            buy_amount = db_config.get('buy_amount', 100000)
            max_daily_buys = db_config.get('max_daily_buys', 8)
        else:
            # 数据库无配置时使用默认值
            score_threshold = 60
            max_hold_days = 10
            stop_loss = -7
            take_profit = 21
            initial_capital = 300000
            buy_amount = 100000
            max_daily_buys = 8
        
        # 温度约束参数（前端传入）
        enable_temp_limit = data.get('enable_temp_limit', 1)
        temp_limit_mode = data.get('temp_limit_mode', 'both')
        
        # 从配置文件读取海龟策略参数
        turtle_params = {}
        if timing_strategy == 'turtle':
            try:
                config_manager = StrategyConfigManager()
                turtle_config = config_manager.get_strategy_config('TurtleStrategy')
                turtle_params = turtle_config.get('params', {})
                logger.info(f"从配置文件读取海龟策略参数: n_entry={turtle_params.get('n_entry')}, "
                           f"n_exit={turtle_params.get('n_exit')}, atr_period={turtle_params.get('atr_period')}")
            except Exception as e:
                logger.warning(f"读取海龟策略配置失败，使用默认值: {str(e)}")
                turtle_params = {
                    'n_entry': 20, 'n_exit': 10, 'atr_period': 20,
                    'entry_atr': 0.02, 'add_atr': 0.5, 'exit_atr': 2.0, 'base_position_amount': 20000
                }
        
        # 验证参数
        if not strategy_name or not start_date or not end_date:
            return jsonify({
                'success': False,
                'message': '缺少必要的执行条件（策略名称、开始日期、结束日期）',
                'data': None
            }), 400
        
        # 构建回测配置
        config = {
            'config_name': data.get('config_name', '默认配置'),
            'score_threshold': score_threshold,
            'hold_period': max_hold_days,
            'stop_loss': stop_loss,
            'take_profit': take_profit,
            'initial_capital': initial_capital,
            'buy_amount': buy_amount,
            'max_daily_buys': max_daily_buys,
            'timing_strategy': timing_strategy,
            'timing_params': timing_params,  # 择时策略参数
            'support_level_method': support_level_method,
            'buy_point_lower': -1,
            'buy_point_upper': 3,
            'start_date': start_date,
            'end_date': end_date,
            # 温度约束参数
            'enable_temp_limit': enable_temp_limit,
            'temp_limit_mode': temp_limit_mode,
            # 海龟策略参数（从配置文件读取）
            'n_entry': turtle_params.get('n_entry'),
            'n_exit': turtle_params.get('n_exit'),
            'atr_period': turtle_params.get('atr_period'),
            'entry_atr': turtle_params.get('entry_atr'),
            'add_atr': turtle_params.get('add_atr'),
            'exit_atr': turtle_params.get('exit_atr'),
            'base_position_amount': turtle_params.get('base_position_amount')
        }
        
        # 使用原有的回测引擎
        logger.info("使用原有回测引擎")
        engine = BacktestEngine()
        
        # 运行回测（使用英文策略名称）
        result = engine.run_backtest(english_strategy_name, config)
        
        # 构建保存到数据库的结果格式
        # 计算final_capital
        final_capital = config.get('initial_capital', 300000)
        if 'capital_history' in result and result['capital_history']:
            final_capital = result['capital_history'][-1]
        
        # 使用中文策略名称保存到数据库，每次都创建新记录
        save_result = {
            'strategy_name': strategy_name,  # 保存中文策略名称
            'support_level_method': timing_strategy,  # 保存择时策略
            'backtest_name': f"{strategy_name}_{start_date}_{end_date}",
            'start_date': start_date,
            'end_date': end_date,
            'total_trades': result.get('performance', {}).get('total_trades', 0),
            'win_trades': result.get('performance', {}).get('win_trades', 0),
            'loss_trades': result.get('performance', {}).get('loss_trades', 0),
            'win_rate': result.get('performance', {}).get('win_rate', 0),
            'avg_return': result.get('performance', {}).get('avg_return', 0),
            'total_return': result.get('performance', {}).get('total_return', 0),
            'max_return': result.get('performance', {}).get('max_return', 0),
            'min_return': result.get('performance', {}).get('min_return', 0),
            'profit_factor': result.get('performance', {}).get('profit_factor', 0),
            'profit_loss_ratio': result.get('performance', {}).get('profit_loss_ratio', 0),
            'max_drawdown': result.get('performance', {}).get('max_drawdown', 0),
            'sharpe_ratio': result.get('performance', {}).get('sharpe_ratio', 0),
            'initial_capital': config.get('initial_capital', 300000),
            'final_capital': final_capital
        }
        
        # 每次都创建新记录，不覆盖已有的回测结果
        logger.info(f"开始保存回测结果，数据: {save_result}")
        result_id = backtest_dao.save_result(save_result)
        logger.info(f"保存回测结果完成，result_id: {result_id}")
        
        # 保存交易记录
        if 'trades' in result:
            trades = result['trades']
            for trade in trades:
                trade['result_id'] = result_id
                # 确保交易记录包含所有必要字段
                trade.setdefault('stock_code', '')
                trade.setdefault('stock_name', '')
                # selection_date 是必填字段，不能为空，使用 buy_date 作为默认值
                if not trade.get('selection_date'):
                    trade['selection_date'] = trade.get('buy_date', start_date)
                trade.setdefault('buy_date', '')
                trade.setdefault('buy_price', 0)
                trade.setdefault('sell_date', '')
                trade.setdefault('sell_price', 0)
                trade.setdefault('buy_amount', 0)
                trade.setdefault('sell_amount', 0)
                trade.setdefault('profit', trade.get('profit_loss', 0))
                trade.setdefault('profit_rate', trade.get('return_rate', 0))
                trade.setdefault('trade_type', 'normal')
            backtest_dao.save_trades_batch(trades)
        
        # 保存收益曲线数据
        if 'capital_history' in result and 'dates' in result:
            capital_history = result['capital_history']
            dates = result['dates']
            equity_curve = []
            initial_capital = config.get('initial_capital', 300000)
            
            # capital_history 是数字列表，每个元素是总资产值
            # dates 是对应的日期列表
            # 注意：capital_history[0] 是初始资金，capital_history[1:] 对应 dates 中的日期
            
            # 添加初始数据点（初始日期、初始资金、收益率0）
            if capital_history:
                equity_curve.append({
                    'date': config.get('start_date', ''),
                    'capital': capital_history[0],
                    'return_rate': 0.0
                })
            
            # 添加后续数据点（确保日期和资金对应正确）
            # capital_history[1:] 对应 dates 中的所有日期
            for i, date_value in enumerate(dates):
                try:
                    # capital_history 中的索引是 i+1（因为第一个元素是初始资金）
                    if i + 1 < len(capital_history):
                        capital_value = capital_history[i + 1]
                    else:
                        # 如果索引超出范围，使用最后一个值
                        capital_value = capital_history[-1] if capital_history else 0
                    
                    # 计算收益率：(当日总资产 / 初始资金 - 1) * 100
                    if capital_value > 0 and initial_capital > 0:
                        return_rate = ((capital_value / initial_capital) - 1) * 100
                    else:
                        return_rate = 0.0
                    
                    # 格式化日期
                    if isinstance(date_value, str):
                        date_str = date_value
                    else:
                        date_str = str(date_value)
                    
                    equity_curve.append({
                        'date': date_str,
                        'capital': capital_value,
                        'return_rate': return_rate
                    })
                except Exception as e:
                    logger.warning(f"处理资金历史记录时出错: {str(e)}")
                    continue
            
            logger.info(f"生成收益曲线数据点数: {len(equity_curve)}")
            backtest_dao.save_equity_curve(result_id, equity_curve)
        
        # 获取完整的回测结果（包括equity_curve）
        full_result = backtest_dao.get_result_by_id(result_id)
        trades = backtest_dao.get_trades_by_result(result_id)
        equity_curve_data = backtest_dao.get_equity_curve(result_id)
        
        # 处理结果中的Infinity值
        def handle_infinity(value):
            if isinstance(value, (int, float)):
                if value == float('inf') or value == float('-inf') or value != value:
                    return None
            return value
        
        # 处理结果中的Infinity值
        processed_result = {}
        for key, value in full_result.items():
            processed_result[key] = handle_infinity(value)
        
        # 处理交易记录中的Infinity值
        processed_trades = []
        for trade in trades:
            processed_trade = {}
            for key, value in trade.items():
                processed_trade[key] = handle_infinity(value)
            processed_trades.append(processed_trade)
        
        # 处理收益曲线中的Infinity值
        processed_equity_curve = []
        for item in equity_curve_data:
            processed_item = {}
            for key, value in item.items():
                processed_item[key] = handle_infinity(value)
            processed_equity_curve.append(processed_item)
        
        # 添加交易记录和收益曲线到结果中
        processed_result['trades'] = processed_trades
        processed_result['equity_curve'] = processed_equity_curve
        
        return jsonify({
            'success': True,
            'message': '回测运行成功',
            'data': processed_result
        }), 200
    
    except Exception as e:
        logger.error(f"运行回测失败: {str(e)}")
        return jsonify({
            'success': False,
            'message': f'运行回测失败: {str(e)}',
            'data': None
        }), 500





@trading_bp.route('/backtest/results', methods=['GET'])
def get_backtest_results():
    """
    获取回测结果列表接口
    
    返回:
        {
            "success": true/false,
            "message": "成功或错误信息",
            "data": {
                "results": [
                    {
                        "id": 1,
                        "strategy_name": "多方炮策略",
                        "config_id": 1,
                        "start_date": "2024-01-01",
                        "end_date": "2024-06-30",
                        "initial_capital": 300000,
                        "final_capital": 331500,
                        "total_return": 10.5,
                        "win_rate": 65.0,
                        "avg_win": 8.2,
                        "avg_loss": -4.1,
                        "profit_factor": 2.0,
                        "max_drawdown": 8.2,
                        "sharpe_ratio": 1.2,
                        "volatility": 15.0,
                        "sortino_ratio": 1.5,
                        "total_trades": 20,
                        "winning_trades": 13,
                        "losing_trades": 7,
                        "holding_period": 10,
                        "created_at": "2024-01-01 12:00:00"
                    }
                ],
                "total_count": 1
            }
        }
    """
    try:
        # 获取查询参数
        strategy_name = request.args.get('strategy')
        created_date = request.args.get('created_date')
        created_start = request.args.get('created_start')
        created_end = request.args.get('created_end')
        
        # 调用DAO获取所有结果（支持按策略名称和创建时间筛选）
        results = backtest_dao.get_all_results(strategy_name, created_date, created_start, created_end)
        
        # 处理结果中的Infinity值，将其转换为null
        def handle_infinity(value):
            if isinstance(value, (int, float)):
                if value == float('inf') or value == float('-inf') or value != value:  # 处理Infinity和NaN
                    return None
            return value
        
        # 处理每个结果
        processed_results = []
        for result in results:
            processed_result = {}
            for key, value in result.items():
                processed_result[key] = handle_infinity(value)
            
            # 确保返回的字段名与前端期望的一致
            processed_result['winning_trades'] = processed_result.get('win_trades', 0)
            processed_result['losing_trades'] = processed_result.get('loss_trades', 0)
            processed_result['profit_loss_ratio'] = processed_result.get('profit_loss_ratio', 0)
            processed_result['avg_hold_days'] = processed_result.get('avg_hold_days', 0)
            processed_result['volatility'] = processed_result.get('volatility', 0)
            processed_result['sortino_ratio'] = processed_result.get('sortino_ratio', 0)
            
            # 确保初始资金和最终资金不为0
            if processed_result.get('initial_capital', 0) == 0:
                processed_result['initial_capital'] = 300000
            if processed_result.get('final_capital', 0) == 0:
                # 根据总收益率计算最终资金
                total_return = processed_result.get('total_return', 0)
                initial_capital = processed_result.get('initial_capital', 300000)
                processed_result['final_capital'] = initial_capital * (1 + total_return / 100)
            
            processed_results.append(processed_result)
        
        return jsonify({
            'success': True,
            'message': '获取回测结果列表成功',
            'data': {
                'results': processed_results,
                'total_count': len(processed_results)
            }
        }), 200
    
    except Exception as e:
        logger.error(f"获取回测结果列表失败: {str(e)}")
        return jsonify({
            'success': False,
            'message': f'获取回测结果列表失败: {str(e)}',
            'data': None
        }), 500


@trading_bp.route('/backtest/results/<int:result_id>', methods=['GET'])
def get_backtest_result(result_id):
    """
    获取单个回测结果详情接口
    
    参数:
        result_id: 结果ID (路径参数)
    
    返回:
        {
            "success": true/false,
            "message": "成功或错误信息",
            "data": {
                "id": 1,
                "strategy_name": "多方炮策略",
                "config_id": 1,
                "start_date": "2024-01-01",
                "end_date": "2024-06-30",
                "initial_capital": 300000,
                "final_capital": 331500,
                "total_return": 10.5,
                "win_rate": 65.0,
                "avg_win": 8.2,
                "avg_loss": -4.1,
                "profit_factor": 2.0,
                "max_drawdown": 8.2,
                "sharpe_ratio": 1.2,
                "volatility": 15.0,
                "sortino_ratio": 1.5,
                "total_trades": 20,
                "winning_trades": 13,
                "losing_trades": 7,
                "holding_period": 10,
                "created_at": "2024-01-01 12:00:00",
                "trades": [...],
                "equity_curve": [...]
            }
        }
    """
    try:
        # 调用DAO获取结果详情
        result = backtest_dao.get_result_by_id(result_id)
        
        if not result:
            return jsonify({
                'success': False,
                'message': '结果不存在',
                'data': None
            }), 404
        
        # 获取交易记录
        trades = backtest_dao.get_trades_by_result(result_id)
        
        # 从数据库获取收益曲线数据
        equity_curve = backtest_dao.get_equity_curve(result_id)
        
        # 转换交易记录的日期格式
        for trade in trades:
            if trade.get('buy_date'):
                trade['buy_date'] = str(trade['buy_date'])
            if trade.get('sell_date'):
                trade['sell_date'] = str(trade['sell_date'])
        
        # 处理结果中的Infinity值，将其转换为null
        def handle_infinity(value):
            if isinstance(value, (int, float)):
                if value == float('inf') or value == float('-inf') or value != value:  # 处理Infinity和NaN
                    return None
            return value
        
        # 处理结果中的Infinity值
        processed_result = {}
        for key, value in result.items():
            processed_result[key] = handle_infinity(value)
        
        # 确保返回的字段名与前端期望的一致
        processed_result['winning_trades'] = processed_result.get('win_trades', 0)
        processed_result['losing_trades'] = processed_result.get('loss_trades', 0)
        processed_result['profit_loss_ratio'] = processed_result.get('profit_loss_ratio', 0)
        processed_result['avg_hold_days'] = processed_result.get('avg_hold_days', 0)
        processed_result['volatility'] = processed_result.get('volatility', 0)
        processed_result['sortino_ratio'] = processed_result.get('sortino_ratio', 0)
        
        # 确保初始资金和最终资金不为0
        if processed_result.get('initial_capital', 0) == 0:
            processed_result['initial_capital'] = 300000
        if processed_result.get('final_capital', 0) == 0:
            # 根据总收益率计算最终资金
            total_return = processed_result.get('total_return', 0)
            initial_capital = processed_result.get('initial_capital', 300000)
            processed_result['final_capital'] = initial_capital * (1 + total_return / 100)
        
        # 处理交易记录中的Infinity值
        processed_trades = []
        for trade in trades:
            processed_trade = {}
            for key, value in trade.items():
                processed_trade[key] = handle_infinity(value)
            processed_trades.append(processed_trade)
        
        # 处理收益曲线中的Infinity值
        processed_equity_curve = []
        for item in equity_curve:
            processed_item = {}
            for key, value in item.items():
                processed_item[key] = handle_infinity(value)
            processed_equity_curve.append(processed_item)
        
        # 添加交易记录和收益曲线到结果中
        processed_result['trades'] = processed_trades
        processed_result['equity_curve'] = processed_equity_curve
        
        return jsonify({
            'success': True,
            'message': '获取回测结果详情成功',
            'data': processed_result
        }), 200
    
    except Exception as e:
        logger.error(f"获取回测结果详情失败: {str(e)}")
        return jsonify({
            'success': False,
            'message': f'获取回测结果详情失败: {str(e)}',
            'data': None
        }), 500


@trading_bp.route('/backtest/results/<int:result_id>/trades', methods=['GET'])
def get_backtest_trades(result_id):
    """
    获取回测交易记录接口
    
    参数:
        result_id: 结果ID (路径参数)
    
    返回:
        {
            "success": true/false,
            "message": "成功或错误信息",
            "data": {
                "trades": [
                    {
                        "id": 1,
                        "result_id": 1,
                        "stock_code": "000001",
                        "stock_name": "平安银行",
                        "buy_date": "2024-01-02",
                        "buy_price": 10.0,
                        "sell_date": "2024-01-12",
                        "sell_price": 10.8,
                        "buy_amount": 100000,
                        "sell_amount": 108000,
                        "profit": 8000,
                        "profit_rate": 8.0,
                        "trade_type": "normal"
                    }
                ],
                "total_count": 1
            }
        }
    """
    try:
        # 调用DAO获取交易记录
        trades = backtest_dao.get_trades_by_result_id(result_id)
        
        return jsonify({
            'success': True,
            'message': '获取回测交易记录成功',
            'data': {
                'trades': trades,
                'total_count': len(trades)
            }
        }), 200
    
    except Exception as e:
        logger.error(f"获取回测交易记录失败: {str(e)}")
        return jsonify({
            'success': False,
            'message': f'获取回测交易记录失败: {str(e)}',
            'data': None
        }), 500


@trading_bp.route('/backtest/results/<int:result_id>/export', methods=['GET'])
def export_backtest_result(result_id):
    """
    导出回测结果为Excel接口
    
    参数:
        result_id: 回测结果ID (路径参数)
    
    返回:
        Excel文件下载，包含三个Sheet：
        1. 基本信息 - 回测概览
        2. 收益曲线 - 每日资金和收益率
        3. 交易明细 - 所有交易记录
    """
    try:
        from io import BytesIO
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
        
        # 获取回测结果
        result = backtest_dao.get_result_by_id(result_id)
        if not result:
            return jsonify({
                'success': False,
                'message': '回测结果不存在',
                'data': None
            }), 404
        
        # 获取交易记录
        trades = backtest_dao.get_trades_by_result(result_id)
        
        # 获取收益曲线
        equity_curve = backtest_dao.get_equity_curve(result_id)
        
        # 创建工作簿
        wb = openpyxl.Workbook()
        
        # 定义样式
        header_fill = PatternFill(start_color="1e3a8a", end_color="1e3a8a", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        sub_header_fill = PatternFill(start_color="dbeafe", end_color="dbeafe", fill_type="solid")
        sub_header_font = Font(bold=True, color="1e3a8a")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # ========== Sheet 1: 基本信息 ==========
        ws1 = wb.active
        ws1.title = "基本信息"
        
        # 设置列宽
        ws1.column_dimensions['A'].width = 25
        ws1.column_dimensions['B'].width = 20
        ws1.column_dimensions['C'].width = 25
        ws1.column_dimensions['D'].width = 20
        
        row = 1
        # 标题
        ws1.cell(row=row, column=1, value="回测结果基本信息").font = Font(bold=True, size=14)
        ws1.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        row += 2
        
        # 策略信息
        ws1.cell(row=row, column=1, value="策略信息").fill = sub_header_fill
        ws1.cell(row=row, column=1).font = sub_header_font
        ws1.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        row += 1
        
        info_rows = [
            ("策略名称", result.get('strategy_name', '')),
            ("回测期间", f"{result.get('start_date', '')} 至 {result.get('end_date', '')}"),
            ("择时策略", result.get('timing_strategy', '')),
            ("支撑位计算方法", result.get('support_level_method', '')),
        ]
        for label, value in info_rows:
            ws1.cell(row=row, column=1, value=label).border = border
            ws1.cell(row=row, column=2, value=value).border = border
            ws1.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
            row += 1
        
        row += 1
        # 资金信息
        ws1.cell(row=row, column=1, value="资金信息").fill = sub_header_fill
        ws1.cell(row=row, column=1).font = sub_header_font
        ws1.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        row += 1
        
        capital_rows = [
            ("初始资金", f"¥{result.get('initial_capital', 0):,.2f}"),
            ("最终资金", f"¥{result.get('final_capital', 0):,.2f}"),
            ("总收益率", f"{result.get('total_return', 0):.2f}%"),
            ("最大回撤", f"{result.get('max_drawdown', 0):.2f}%"),
        ]
        for label, value in capital_rows:
            ws1.cell(row=row, column=1, value=label).border = border
            ws1.cell(row=row, column=2, value=value).border = border
            ws1.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
            row += 1
        
        row += 1
        # 交易统计
        ws1.cell(row=row, column=1, value="交易统计").fill = sub_header_fill
        ws1.cell(row=row, column=1).font = sub_header_font
        ws1.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        row += 1
        
        trade_rows = [
            ("总交易次数", result.get('total_trades', 0)),
            ("盈利交易", result.get('win_trades', 0)),
            ("亏损交易", result.get('loss_trades', 0)),
            ("胜率", f"{result.get('win_rate', 0):.2f}%"),
            ("盈亏比", f"{result.get('profit_loss_ratio', 0):.2f}"),
            ("平均收益率", f"{result.get('avg_return', 0):.2f}%"),
        ]
        for label, value in trade_rows:
            ws1.cell(row=row, column=1, value=label).border = border
            ws1.cell(row=row, column=2, value=value).border = border
            ws1.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
            row += 1
        
        row += 1
        # 风险指标
        ws1.cell(row=row, column=1, value="风险指标").fill = sub_header_fill
        ws1.cell(row=row, column=1).font = sub_header_font
        ws1.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        row += 1
        
        risk_rows = [
            ("夏普比率", f"{result.get('sharpe_ratio', 0):.2f}"),
            ("索提诺比率", f"{result.get('sortino_ratio', 0):.2f}"),
            ("波动率", f"{result.get('volatility', 0):.2f}%"),
            ("最大单笔收益", f"{result.get('max_return', 0):.2f}%"),
            ("最小单笔收益", f"{result.get('min_return', 0):.2f}%"),
        ]
        for label, value in risk_rows:
            ws1.cell(row=row, column=1, value=label).border = border
            ws1.cell(row=row, column=2, value=value).border = border
            ws1.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
            row += 1
        
        # ========== Sheet 2: 收益曲线 ==========
        ws2 = wb.create_sheet(title="收益曲线")
        
        ws2.column_dimensions['A'].width = 15
        ws2.column_dimensions['B'].width = 18
        ws2.column_dimensions['C'].width = 15
        
        # 表头
        row = 1
        headers = ['日期', '资金', '收益率(%)']
        for col, header in enumerate(headers, 1):
            cell = ws2.cell(row=row, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        
        # 数据
        row = 2
        for item in equity_curve:
            ws2.cell(row=row, column=1, value=item.get('date', '')).border = border
            ws2.cell(row=row, column=2, value=item.get('capital', 0)).border = border
            ws2.cell(row=row, column=2).number_format = '#,##0.00'
            ws2.cell(row=row, column=3, value=item.get('return_rate', 0)).border = border
            ws2.cell(row=row, column=3).number_format = '0.00'
            row += 1
        
        # ========== Sheet 3: 交易明细 ==========
        ws3 = wb.create_sheet(title="交易明细")
        
        ws3.column_dimensions['A'].width = 12
        ws3.column_dimensions['B'].width = 12
        ws3.column_dimensions['C'].width = 15
        ws3.column_dimensions['D'].width = 12
        ws3.column_dimensions['E'].width = 12
        ws3.column_dimensions['F'].width = 12
        ws3.column_dimensions['G'].width = 12
        ws3.column_dimensions['H'].width = 10
        ws3.column_dimensions['I'].width = 12
        
        # 表头
        row = 1
        headers = ['股票代码', '股票名称', '买入日期', '买入价格', '买入金额', '卖出日期', '卖出价格', '收益率(%)', '卖出类型']
        for col, header in enumerate(headers, 1):
            cell = ws3.cell(row=row, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        
        # 数据
        row = 2
        for trade in trades:
            ws3.cell(row=row, column=1, value=trade.get('stock_code', '')).border = border
            ws3.cell(row=row, column=2, value=trade.get('stock_name', '')).border = border
            ws3.cell(row=row, column=3, value=trade.get('buy_date', '')).border = border
            ws3.cell(row=row, column=4, value=trade.get('buy_price', 0)).border = border
            ws3.cell(row=row, column=4).number_format = '0.00'
            ws3.cell(row=row, column=5, value=trade.get('buy_amount', 0)).border = border
            ws3.cell(row=row, column=5).number_format = '#,##0.00'
            ws3.cell(row=row, column=6, value=trade.get('sell_date', '')).border = border
            ws3.cell(row=row, column=7, value=trade.get('sell_price', 0)).border = border
            ws3.cell(row=row, column=7).number_format = '0.00'
            ws3.cell(row=row, column=8, value=trade.get('return_rate', 0)).border = border
            ws3.cell(row=row, column=8).number_format = '0.00'
            ws3.cell(row=row, column=9, value=trade.get('sell_type', '')).border = border
            row += 1
        
        # 保存文件
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        # 生成文件名
        strategy_name = result.get('strategy_name', '回测结果')
        start_date = result.get('start_date', '').replace('-', '')
        end_date = result.get('end_date', '').replace('-', '')
        filename = f"回测报告_{strategy_name}_{start_date}_{end_date}.xlsx"
        
        # 创建响应
        from flask import make_response
        from urllib.parse import quote
        response = make_response(output.getvalue())
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        encoded_filename = quote(filename)
        response.headers['Content-Disposition'] = f'attachment; filename="{encoded_filename}"'
        return response
        
    except Exception as e:
        logger.error(f"导出回测结果失败: {str(e)}")
        import traceback
        logger.error(traceback.format_exc())
        return jsonify({
            'success': False,
            'message': f'导出回测结果失败: {str(e)}',
            'data': None
        }), 500


@trading_bp.route('/backtest/strategies', methods=['GET'])
def get_strategies():
    """
    获取可用策略列表接口
    
    返回:
        {
            "success": true/false,
            "message": "成功或错误信息",
            "data": {
                "strategies": [
                    {
                        "name": "连阳回调策略",
                        "display_name": "连阳回调策略"
                    }
                ]
            }
        }
    """
    try:
        # 从策略注册表获取策略列表
        from strategy.strategy_registry import get_registry
        registry = get_registry("config/strategy_params.yaml")
        registry.auto_register_from_directory("strategy")
        
        strategies = []
        for strategy_name in registry.list_strategies():
            # 获取策略对象
            strategy = registry.get_strategy(strategy_name)
            if strategy:
                # 使用策略对象的name属性（中文名称）
                chinese_name = strategy.name
                strategies.append({
                    'name': chinese_name,  # 中文名称
                    'display_name': chinese_name  # 中文名称
                })
        
        return jsonify({
            'success': True,
            'message': '获取策略列表成功',
            'data': {
                'strategies': strategies
            }
        }), 200
    
    except Exception as e:
        logger.error(f"获取策略列表失败: {str(e)}")
        import traceback
        logger.error(traceback.format_exc())
        return jsonify({
            'success': False,
            'message': f'获取策略列表失败: {str(e)}',
            'data': None
        }), 500


# ==================== KHunter 狩猎场相关接口 ====================

# 创建 KHunter 蓝图
khunter_bp = Blueprint('khunter', __name__, url_prefix='/api/khunter')

# 初始化 KHunter 相关组件
from trading.khunter_api import KHunterAPI
from trading.khunter_data_processor import KHunterDataProcessor
from trading.khunter_dao import KHunterDAO
from trading.khunter_support_calculator import KHunterSupportCalculator
from trading.khunter_buy_point_judge import KHunterBuyPointJudge
from utils.strategy_config_manager import StrategyConfigManager

# 1. 初始化策略配置管理器
strategy_config_manager = StrategyConfigManager()

# 2. 初始化 KHunter DAO
khunter_dao = KHunterDAO(db_manager)

# 3. 初始化支撑位计算器（传入配置管理器）
khunter_support_calculator = KHunterSupportCalculator(db_manager, strategy_config_manager)

# 4. 初始化买点判断器
khunter_buy_point_judge = KHunterBuyPointJudge(db_manager)

# 5. 初始化数据处理器
khunter_data_processor = KHunterDataProcessor(db_manager, khunter_support_calculator, khunter_buy_point_judge)

# 6. 初始化 API
khunter_api = KHunterAPI(db_manager, khunter_data_processor, khunter_dao)


@khunter_bp.route('/calculate', methods=['POST'])
def khunter_calculate():
    """
    计算狩猎场数据接口
    
    请求体:
        {
            "hunting_date": "2024-04-15",
            "tracking_days": 10
        }
    
    返回:
        {
            "success": true/false,
            "message": "成功或错误信息",
            "data": {
                "hunting_date": "2024-04-15",
                "tracking_days": 10,
                "from_cache": false,
                "total_count": 8,
                "calculation_time": 2.5,
                "results": [...]
            }
        }
    """
    try:
        # 1. 获取请求数据
        data = request.get_json() or {}
        hunting_date = data.get('hunting_date')
        tracking_days = data.get('tracking_days', 10)
        # 如果前端传递空字符串或null，使用默认值'support'
        timing_strategy = data.get('timing_strategy') or 'support'
        
        # 1a. 记录请求参数，便于调试前端传参
        logger.info(f"狩猎场计算请求: hunting_date={hunting_date}, tracking_days={tracking_days}, timing_strategy={timing_strategy}")
        
        # 2. 调用 API 计算
        result = khunter_api.calculate(hunting_date, tracking_days, timing_strategy)
        
        # 3. 返回结果
        return jsonify(result), 200 if result.get('success') else 400
    
    except Exception as e:
        logger.error(f"计算狩猎场数据失败: {str(e)}")
        return jsonify({
            'success': False,
            'message': f'计算失败：{str(e)}',
            'data': None
        }), 500


@khunter_bp.route('/save', methods=['POST'])
def khunter_save():
    """
    保存计算结果接口
    
    请求体:
        {
            "hunting_date": "2024-04-15",
            "tracking_days": 10
        }
    
    返回:
        {
            "success": true/false,
            "message": "成功或错误信息",
            "data": {
                "saved_count": 8,
                "hunting_date": "2024-04-15"
            }
        }
    """
    try:
        # 1. 获取请求数据
        data = request.get_json() or {}
        hunting_date = data.get('hunting_date')
        tracking_days = data.get('tracking_days', 10)
        timing_strategy = data.get('timing_strategy', 'support')
        
        # 2. 调用 API 保存
        result = khunter_api.save(hunting_date, tracking_days, timing_strategy)
        
        # 3. 返回结果
        return jsonify(result), 200 if result.get('success') else 400
    
    except Exception as e:
        logger.error(f"保存狩猎场数据失败: {str(e)}")
        return jsonify({
            'success': False,
            'message': f'保存失败：{str(e)}',
            'data': None
        }), 500


@khunter_bp.route('/query', methods=['GET'])
def khunter_query():
    """
    查询狩猎场数据接口
    
    查询参数:
        hunting_date: 狩猎日期 (必填)
    
    返回:
        {
            "success": true/false,
            "message": "成功或错误信息",
            "data": {
                "total_count": 25,
                "results": [...]
            }
        }
    """
    try:
        # 1. 获取查询参数
        hunting_date = request.args.get('hunting_date')
        timing_strategy = request.args.get('timing_strategy')
        
        # 2. 调用 API 查询
        result = khunter_api.query(hunting_date, timing_strategy)
        
        # 3. 返回结果
        return jsonify(result), 200 if result.get('success') else 400
    
    except Exception as e:
        logger.error(f"查询狩猎场数据失败: {str(e)}")
        return jsonify({
            'success': False,
            'message': f'查询失败：{str(e)}',
            'data': None
        }), 500


@khunter_bp.route('/check-cache', methods=['GET'])
def khunter_check_cache():
    """
    检查缓存接口
    
    查询参数:
        hunting_date: 狩猎日期 (必填)
    
    返回:
        {
            "success": true/false,
            "message": "成功或错误信息",
            "data": {
                "has_cache": true,
                "record_count": 8,
                "hunting_date": "2024-04-15"
            }
        }
    """
    try:
        # 1. 获取查询参数
        hunting_date = request.args.get('hunting_date')
        timing_strategy = request.args.get('timing_strategy')
        
        # 2. 调用 API 检查缓存
        result = khunter_api.check_cache(hunting_date, timing_strategy)
        
        # 3. 返回结果
        return jsonify(result), 200 if result.get('success') else 400
    
    except Exception as e:
        logger.error(f"检查狩猎场缓存失败: {str(e)}")
        return jsonify({
            'success': False,
            'message': f'检查失败：{str(e)}',
            'data': None
        }), 500


@khunter_bp.route('/track', methods=['GET'])
def khunter_track():
    """
    狩猎跟踪接口 - 获取指定日期的全部狩猎数据
    
    查询参数:
        hunting_date: 狩猎日期 (必填)
    
    返回:
        {
            "success": true/false,
            "message": "成功或错误信息",
            "data": [
                {
                    "rank_position": 1,
                    "stock_code": "000001",
                    "stock_name": "平安银行",
                    "score": 85.5,
                    "industry": "银行",
                    "sector": "金融",
                    "selection_price": 10.5,
                    "current_price": 10.8,
                    "current_yield": 2.86,
                    "highest_price": 11.2,
                    "highest_yield": 6.67
                }
            ]
        }
    """
    try:
        # 1. 获取查询参数
        hunting_date = request.args.get('hunting_date')
        
        # 2. 调用 API 获取跟踪数据（获取全部数据）
        result = khunter_api.track(hunting_date)
        
        # 3. 返回结果
        return jsonify(result), 200 if result.get('success') else 400
    
    except Exception as e:
        logger.error(f"狩猎跟踪失败: {str(e)}")
        return jsonify({
            'success': False,
            'message': f'跟踪失败：{str(e)}',
            'data': None
        }), 500


@khunter_bp.route('/latest_kline_date', methods=['GET'])
def get_latest_kline_date():
    """
    获取最后一根K线日期接口
    
    返回:
        {
            "success": true/false,
            "message": "成功或错误信息",
            "data": {
                "latest_date": "2026-04-20"
            }
        }
    """
    try:
        result = khunter_api.get_latest_kline_date()
        return jsonify(result), 200 if result.get('success') else 400
    
    except Exception as e:
        logger.error(f"获取最后一根K线日期失败: {str(e)}")
        return jsonify({
            'success': False,
            'message': f'获取失败：{str(e)}',
            'data': None
        }), 500


# ==================== 交易计划相关接口 ====================

from trading.trading_plan_generator import TradingPlanGenerator
from trading.trading_plan_dao import TradingPlanDAO

trading_plan_dao = TradingPlanDAO(db_manager)
trading_plan_generator = TradingPlanGenerator(db_manager, khunter_dao, trading_plan_dao)


@khunter_bp.route('/generate_plan', methods=['POST'])
def generate_trading_plan():
    """
    生成交易计划接口

    请求体:
        {
            "hunting_date": "2024-04-15"
        }

    返回:
        {
            "success": true/false,
            "message": "成功或错误信息",
            "data": {
                "plan_date": "2024-04-16",
                "hunting_date": "2024-04-15",
                "total_count": 8,
                "plans": [...]
            }
        }
    """
    try:
        data = request.get_json() or {}
        hunting_date = data.get('hunting_date')

        if not hunting_date:
            return jsonify({
                'success': False,
                'message': '缺少必填参数: hunting_date',
                'data': None
            }), 400

        result = trading_plan_generator.generate(hunting_date)
        return jsonify({
            'success': True,
            'message': '交易计划生成成功',
            'data': result
        }), 200

    except Exception as e:
        logger.error(f"生成交易计划失败: {str(e)}")
        return jsonify({
            'success': False,
            'message': f'生成交易计划失败：{str(e)}',
            'data': None
        }), 500


@khunter_bp.route('/export_plan', methods=['POST'])
def export_trading_plan():
    """
    导出交易计划Excel接口

    请求体:
        {
            "hunting_date": "2024-04-15"
        }

    返回:
        Excel文件下载
    """
    try:
        data = request.get_json() or {}
        hunting_date = data.get('hunting_date')

        if not hunting_date:
            return jsonify({
                'success': False,
                'message': '缺少必填参数: hunting_date',
                'data': None
            }), 400

        # 直接生成交易计划，不保存到数据库
        result = trading_plan_generator.generate(hunting_date)
        plans = result.get('plans', [])
        
        if not plans:
            return jsonify({
                'success': False,
                'message': '没有找到对应的交易计划',
                'data': None
            }), 404

        from io import BytesIO
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "交易计划"

        headers = ['序号', '股票代码', '股票名称', '当前价格', '买入价格区间', '仓位(%)', '止损', '止盈', '择时策略']
        header_fill = PatternFill(start_color="1e3a8a", end_color="1e3a8a", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

        for idx, plan in enumerate(plans, 1):
            ws.cell(row=idx+1, column=1, value=idx)
            ws.cell(row=idx+1, column=2, value=plan.get('stock_code', ''))
            ws.cell(row=idx+1, column=3, value=plan.get('stock_name', ''))
            ws.cell(row=idx+1, column=4, value=plan.get('current_price', 0))
            ws.cell(row=idx+1, column=5, value=f"{plan.get('buy_lower_price', 0)}-{plan.get('buy_upper_price', 0)}")
            ws.cell(row=idx+1, column=6, value=plan.get('position_ratio', 0))
            ws.cell(row=idx+1, column=7, value=plan.get('stop_loss_price', 0))
            ws.cell(row=idx+1, column=8, value=plan.get('take_profit_price', 0))
            ws.cell(row=idx+1, column=9, value=f"{plan.get('hold_days', 0)}天")

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        plan_date = result.get('plan_date', '')
        # 使用ASCII兼容的文件名，避免HTTP头编码问题
        filename = f"trading_plan_{plan_date}.xlsx"

        # 创建响应对象并手动设置Content-Disposition头
        from flask import make_response
        from urllib.parse import quote
        response = make_response(output.getvalue())
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        # 对中文文件名进行URL编码
        encoded_filename = quote(filename)
        response.headers['Content-Disposition'] = f'attachment; filename="{encoded_filename}"'
        return response

    except Exception as e:
        logger.error(f"导出交易计划失败: {str(e)}")
        return jsonify({
            'success': False,
            'message': f'导出交易计划失败：{str(e)}',
            'data': None
        }), 500


@khunter_bp.route('/export_selection', methods=['POST'])
def export_selection():
    """
    导出选股结果Excel接口

    请求体:
        {
            "results": {...},      // 选股结果数据
            "selection_date": "2024-04-15",
            "selection_time": "2024-04-15 14:30:00"
        }

    返回:
        Excel文件下载，包含两个Sheet：
        1. 选股结果 - 所有选中的股票
        2. 策略统计 - 各策略选股统计
    """
    try:
        from io import BytesIO
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from urllib.parse import quote
        from flask import make_response

        data = request.get_json() or {}
        results = data.get('results', {})
        selection_date = data.get('selection_date', '')
        selection_time = data.get('selection_time', '')

        # 创建工作簿
        wb = openpyxl.Workbook()

        # 定义样式
        header_fill = PatternFill(start_color="1e3a8a", end_color="1e3a8a", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # ========== Sheet 1: 选股结果 ==========
        ws1 = wb.active
        ws1.title = "选股结果"

        # 设置列宽
        ws1.column_dimensions['A'].width = 12
        ws1.column_dimensions['B'].width = 14
        ws1.column_dimensions['C'].width = 15
        ws1.column_dimensions['D'].width = 25
        ws1.column_dimensions['E'].width = 50
        ws1.column_dimensions['F'].width = 12

        # 表头
        row = 1
        headers = ['序号', '股票代码', '股票名称', '策略名称', '关键日期', '选股理由', '评分']
        for col, header in enumerate(headers, 1):
            cell = ws1.cell(row=row, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border

        # 收集所有股票数据
        row = 2
        all_stocks = []
        strategy_stats = {}  # 策略统计

        for strategy_name, signals in results.items():
            # 跳过特殊字段
            if strategy_name.startswith('_'):
                continue

            if strategy_name not in strategy_stats:
                strategy_stats[strategy_name] = {'count': 0, 'stocks': []}

            if Array.isArray(signals) if 'Array' in dir() else isinstance(signals, list):
                strategy_stats[strategy_name]['count'] = len(signals)
                strategy_stats[strategy_name]['stocks'] = signals

            if isinstance(signals, list):
                strategy_stats[strategy_name]['count'] = len(signals)
                strategy_stats[strategy_name]['stocks'] = signals

                for signal in signals:
                    if not isinstance(signal, dict):
                        continue

                    # 提取关键日期信息
                    key_date_str = ''
                    if signal.get('signals') and isinstance(signal['signals'], list) and len(signal['signals']) > 0:
                        s = signal['signals'][0]
                        if s.get('key_date'):
                            key_type = s.get('key_date_type', '')
                            key_date_str = f"{key_type}: {s['key_date']}" if key_type else s['key_date']

                    # 提取选股理由
                    reasons_str = ''
                    if signal.get('signals') and isinstance(signal['signals'], list) and len(signal['signals']) > 0:
                        s = signal['signals'][0]
                        if s.get('reasons') and isinstance(s['reasons'], list):
                            reasons_str = '; '.join(s['reasons'])

                    all_stocks.append({
                        'code': signal.get('code', ''),
                        'name': signal.get('name', ''),
                        'strategy': strategy_name,
                        'key_date': key_date_str,
                        'reasons': reasons_str,
                        'score': signal.get('score') or signal.get('total_score') or ''
                    })

        # 添加交集股票（被多个策略同时选中的）
        intersection_analysis = results.get('_intersection_analysis', {})
        if isinstance(intersection_analysis, dict) and intersection_analysis.get('by_count'):
            by_count = intersection_analysis.get('by_count', {})
            for count_str, stocks in by_count.items():
                count = int(count_str) if count_str.isdigit() else 0
                if count > 1 and isinstance(stocks, list):
                    for stock in stocks:
                        if not isinstance(stock, dict):
                            continue
                        # 检查是否已存在
                        if not any(s['code'] == stock.get('code') for s in all_stocks):
                            reasons_str = ''
                            if stock.get('reasons') and isinstance(stock['reasons'], list):
                                reasons_str = '; '.join(stock['reasons'])
                            all_stocks.append({
                                'code': stock.get('code', ''),
                                'name': stock.get('name', ''),
                                'strategy': '交集',
                                'key_date': f"被{count}个策略同时选中",
                                'reasons': reasons_str,
                                'score': stock.get('score') or ''
                            })

        # 按评分排序
        all_stocks.sort(key=lambda x: float(x['score']) if x['score'] and str(x['score']).replace('.', '').isdigit() else 0, reverse=True)

        # 写入数据
        for idx, stock in enumerate(all_stocks, 1):
            ws1.cell(row=row, column=1, value=idx).border = border
            ws1.cell(row=row, column=2, value=stock['code']).border = border
            ws1.cell(row=row, column=3, value=stock['name']).border = border
            ws1.cell(row=row, column=4, value=stock['strategy']).border = border
            ws1.cell(row=row, column=5, value=stock['key_date']).border = border
            ws1.cell(row=row, column=6, value=stock['reasons']).border = border
            ws1.cell(row=row, column=7, value=stock['score']).border = border
            row += 1

        # ========== Sheet 2: 策略统计 ==========
        ws2 = wb.create_sheet(title="策略统计")

        ws2.column_dimensions['A'].width = 30
        ws2.column_dimensions['B'].width = 15

        # 表头
        row = 1
        stats_headers = ['策略名称', '选股数量']
        for col, header in enumerate(stats_headers, 1):
            cell = ws2.cell(row=row, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border

        # 添加统计信息
        row = 2
        ws2.cell(row=row, column=1, value='选股日期').border = border
        ws2.cell(row=row, column=2, value=selection_date).border = border
        row += 1
        ws2.cell(row=row, column=1, value='选股时间').border = border
        ws2.cell(row=row, column=2, value=selection_time).border = border
        row += 1
        ws2.cell(row=row, column=1, value='总选股数').border = border
        ws2.cell(row=row, column=2, value=len(all_stocks)).border = border
        row += 2

        # 策略统计
        for strategy_name, stats in sorted(strategy_stats.items(), key=lambda x: x[1]['count'], reverse=True):
            ws2.cell(row=row, column=1, value=strategy_name).border = border
            ws2.cell(row=row, column=2, value=stats['count']).border = border
            row += 1

        # 保存文件
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        # 生成文件名
        date_str = selection_date.replace('-', '') if selection_date else ''
        filename = f"选股结果_{date_str}.xlsx"

        response = make_response(output.getvalue())
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        response.headers['Content-Disposition'] = f'attachment; filename="{quote(filename)}"'
        return response

    except Exception as e:
        logger.error(f"导出选股结果失败: {str(e)}")
        import traceback
        logger.error(traceback.format_exc())
        return jsonify({
            'success': False,
            'message': f'导出选股结果失败：{str(e)}',
            'data': None
        }), 500


# ==================== 执行方案管理接口 ====================

@trading_bp.route('/execution/plans', methods=['GET'])
def get_execution_plans():
    """
    获取所有执行方案列表接口
    
    返回:
        {
            "success": true/false,
            "message": "成功或错误信息",
            "data": {
                "plans": [
                    {
                        "id": "plan_id",
                        "name": "方案名称",
                        "description": "方案描述",
                        "combination_count": 3,
                        "config_ref": "default",
                        "created_at": "2024-01-01T12:00:00",
                        "updated_at": "2024-01-01T12:00:00"
                    }
                ],
                "total_count": 1
            }
        }
    """
    try:
        from trading.strategy_execution_plan import ExecutionPlan
        
        plans = ExecutionPlan.list_plans()
        
        plan_list = []
        for plan in plans:
            plan_list.append({
                'id': plan.id,
                'name': plan.name,
                'description': plan.description,
                'combination_count': len(plan.combinations),
                'config_ref': plan.config_ref,
                'created_at': plan.created_at,
                'updated_at': plan.updated_at
            })
        
        return jsonify({
            'success': True,
            'message': '获取执行方案列表成功',
            'data': {
                'plans': plan_list,
                'total_count': len(plan_list)
            }
        }), 200
    
    except Exception as e:
        logger.error(f"获取执行方案列表失败: {str(e)}")
        return jsonify({
            'success': False,
            'message': f'获取执行方案列表失败: {str(e)}',
            'data': None
        }), 500


@trading_bp.route('/execution/plans/<plan_id>', methods=['GET'])
def get_execution_plan(plan_id):
    """
    获取单个执行方案详情接口
    
    参数:
        plan_id: 方案ID (路径参数)
    
    返回:
        {
            "success": true/false,
            "message": "成功或错误信息",
            "data": {
                "id": "plan_id",
                "name": "方案名称",
                "description": "方案描述",
                "combinations": [
                    {
                        "id": "combo_id",
                        "selection_strategy": "选股策略名称",
                        "timing_strategy": "择时策略名称",
                        "enabled": true
                    }
                ],
                "config_ref": "default",
                "created_at": "2024-01-01T12:00:00",
                "updated_at": "2024-01-01T12:00:00"
            }
        }
    """
    try:
        from trading.strategy_execution_plan import ExecutionPlan
        
        plan = ExecutionPlan.load(plan_id)
        
        if not plan:
            return jsonify({
                'success': False,
                'message': '方案不存在',
                'data': None
            }), 404
        
        combinations = []
        for combo in plan.combinations:
            combinations.append({
                'id': combo.id,
                'selection_strategy': combo.selection_strategy,
                'timing_strategy': combo.timing_strategy,
                'enabled': combo.enabled
            })
        
        return jsonify({
            'success': True,
            'message': '获取执行方案详情成功',
            'data': {
                'id': plan.id,
                'name': plan.name,
                'description': plan.description,
                'combinations': combinations,
                'config_ref': plan.config_ref,
                'created_at': plan.created_at,
                'updated_at': plan.updated_at
            }
        }), 200
    
    except FileNotFoundError:
        return jsonify({
            'success': False,
            'message': '方案不存在',
            'data': None
        }), 404
    except Exception as e:
        logger.error(f"获取执行方案详情失败: {str(e)}")
        return jsonify({
            'success': False,
            'message': f'获取执行方案详情失败: {str(e)}',
            'data': None
        }), 500


@trading_bp.route('/execution/plans', methods=['POST'])
def create_execution_plan():
    """
    创建执行方案接口
    
    请求体:
        {
            "name": "方案名称",
            "description": "方案描述",
            "combinations": [
                {
                    "selection_strategy": "选股策略名称",
                    "timing_strategy": "择时策略名称",
                    "enabled": true
                }
            ],
            "config_ref": "default"
        }
    
    返回:
        {
            "success": true/false,
            "message": "成功或错误信息",
            "data": {
                "plan_id": "plan_id"
            }
        }
    """
    try:
        from trading.strategy_execution_plan import ExecutionPlan, StrategyCombination
        
        data = request.get_json() or {}
        
        if 'name' not in data or not data['name']:
            return jsonify({
                'success': False,
                'message': '方案名称不能为空',
                'data': None
            }), 400
        
        plan = ExecutionPlan(
            name=data['name'],
            description=data.get('description', '')
        )
        
        plan.config_ref = data.get('config_ref', 'default')
        
        # 添加策略组合
        combinations = data.get('combinations', [])
        for combo_data in combinations:
            combo = StrategyCombination(
                selection_strategy=combo_data.get('selection_strategy', ''),
                timing_strategy=combo_data.get('timing_strategy', ''),
                enabled=combo_data.get('enabled', True)
            )
            plan.add_combination(combo)
        
        # 验证并保存
        if not plan.validate():
            return jsonify({
                'success': False,
                'message': '方案验证失败，请确保至少包含一个有效的策略组合',
                'data': None
            }), 400
        
        plan.save()
        
        return jsonify({
            'success': True,
            'message': '创建执行方案成功',
            'data': {
                'plan_id': plan.id
            }
        }), 201
    
    except ValueError as e:
        return jsonify({
            'success': False,
            'message': str(e),
            'data': None
        }), 400
    except Exception as e:
        logger.error(f"创建执行方案失败: {str(e)}")
        return jsonify({
            'success': False,
            'message': f'创建执行方案失败: {str(e)}',
            'data': None
        }), 500


@trading_bp.route('/execution/plans/<plan_id>', methods=['PUT'])
def update_execution_plan(plan_id):
    """
    更新执行方案接口
    
    参数:
        plan_id: 方案ID (路径参数)
    
    请求体:
        {
            "name": "方案名称",
            "description": "方案描述",
            "combinations": [
                {
                    "id": "combo_id",
                    "selection_strategy": "选股策略名称",
                    "timing_strategy": "择时策略名称",
                    "enabled": true
                }
            ],
            "config_ref": "default"
        }
    
    返回:
        {
            "success": true/false,
            "message": "成功或错误信息",
            "data": {
                "plan_id": "plan_id"
            }
        }
    """
    try:
        from trading.strategy_execution_plan import ExecutionPlan, StrategyCombination
        
        data = request.get_json() or {}
        
        plan = ExecutionPlan.load(plan_id)
        
        if not plan:
            return jsonify({
                'success': False,
                'message': '方案不存在',
                'data': None
            }), 404
        
        # 更新基本信息
        if 'name' in data:
            plan.name = data['name']
        if 'description' in data:
            plan.description = data['description']
        if 'config_ref' in data:
            plan.config_ref = data['config_ref']
        
        # 更新策略组合
        if 'combinations' in data:
            plan.combinations = []
            for combo_data in data['combinations']:
                combo = StrategyCombination(
                    combination_id=combo_data.get('id'),
                    selection_strategy=combo_data.get('selection_strategy', ''),
                    timing_strategy=combo_data.get('timing_strategy', ''),
                    enabled=combo_data.get('enabled', True)
                )
                plan.add_combination(combo)
        
        # 验证并保存
        if not plan.validate():
            return jsonify({
                'success': False,
                'message': '方案验证失败',
                'data': None
            }), 400
        
        plan.save()
        
        return jsonify({
            'success': True,
            'message': '更新执行方案成功',
            'data': {
                'plan_id': plan.id
            }
        }), 200
    
    except FileNotFoundError:
        return jsonify({
            'success': False,
            'message': '方案不存在',
            'data': None
        }), 404
    except ValueError as e:
        return jsonify({
            'success': False,
            'message': str(e),
            'data': None
        }), 400
    except Exception as e:
        logger.error(f"更新执行方案失败: {str(e)}")
        return jsonify({
            'success': False,
            'message': f'更新执行方案失败: {str(e)}',
            'data': None
        }), 500


@trading_bp.route('/execution/plans/<plan_id>', methods=['DELETE'])
def delete_execution_plan(plan_id):
    """
    删除执行方案接口
    
    参数:
        plan_id: 方案ID (路径参数)
    
    返回:
        {
            "success": true/false,
            "message": "成功或错误信息",
            "data": {
                "plan_id": "plan_id"
            }
        }
    """
    try:
        from trading.strategy_execution_plan import ExecutionPlan
        
        plan = ExecutionPlan.load(plan_id)
        
        if not plan:
            return jsonify({
                'success': False,
                'message': '方案不存在',
                'data': None
            }), 404
        
        plan.delete()
        
        return jsonify({
            'success': True,
            'message': '删除执行方案成功',
            'data': {
                'plan_id': plan_id
            }
        }), 200
    
    except FileNotFoundError:
        return jsonify({
            'success': False,
            'message': '方案不存在',
            'data': None
        }), 404
    except Exception as e:
        logger.error(f"删除执行方案失败: {str(e)}")
        return jsonify({
            'success': False,
            'message': f'删除执行方案失败: {str(e)}',
            'data': None
        }), 500


@trading_bp.route('/execution/plans/<plan_id>/run', methods=['POST'])
def run_execution_plan(plan_id):
    """
    运行执行方案接口
    
    参数:
        plan_id: 方案ID (路径参数)
    
    请求体:
        {
            "initial_cash": 300000,
            "max_stocks": 8,
            "score_threshold": 60,
            "n_entry": 20,
            "n_exit": 10,
            "atr_period": 20,
            "entry_atr": 0.02,
            "add_atr": 0.5,
            "exit_atr": 2.0,
            "base_position_amount": 20000,
            "turtle_preset": "default"
        }
    
    返回:
        {
            "success": true/false,
            "message": "成功或错误信息",
            "data": {
                "run_date": "2024-01-01",
                "plan_id": "plan_id",
                "plan_name": "方案名称",
                "is_first_run": false,
                "pool_count": 10,
                "total_signals": 5,
                "buy_signals": 3,
                "sell_signals": 2,
                "final_portfolio": {
                    "position_count": 5
                },
                "combination_results": [
                    {
                        "combination_id": "combo_id",
                        "selection_strategy": "选股策略",
                        "timing_strategy": "择时策略",
                        "selected_count": 5,
                        "pool_count_after": 10
                    }
                ]
            }
        }
    """
    try:
        from trading.strategy_execution_plan import ExecutionPlan
        from trading.strategy_runner import StrategyRunner
        
        # 加载方案
        plan = ExecutionPlan.load(plan_id)
        
        if not plan:
            return jsonify({
                'success': False,
                'message': '方案不存在',
                'data': None
            }), 404
        
        # 获取请求配置
        data = request.get_json() or {}
        
        config = {
            'initial_cash': data.get('initial_cash', 300000),
            'max_stocks': data.get('max_stocks', 8),
            'score_threshold': data.get('score_threshold', 60),
            'n_entry': data.get('n_entry'),
            'n_exit': data.get('n_exit'),
            'atr_period': data.get('atr_period'),
            'entry_atr': data.get('entry_atr'),
            'add_atr': data.get('add_atr'),
            'exit_atr': data.get('exit_atr'),
            'base_position_amount': data.get('base_position_amount'),
            'turtle_preset': data.get('turtle_preset')
        }
        
        # 创建策略运行器并执行方案
        runner = StrategyRunner()
        result = runner.run_plan(plan, config)
        
        if result['status'] == 'success':
            return jsonify({
                'success': True,
                'message': result.get('message', '方案执行成功'),
                'data': result.get('data', {})
            }), 200
        else:
            return jsonify({
                'success': False,
                'message': result.get('message', '方案执行失败'),
                'data': None
            }), 500
    
    except FileNotFoundError:
        return jsonify({
            'success': False,
            'message': '方案不存在',
            'data': None
        }), 404
    except Exception as e:
        logger.error(f"运行执行方案失败: {str(e)}")
        import traceback
        logger.error(traceback.format_exc())
        return jsonify({
            'success': False,
            'message': f'运行执行方案失败: {str(e)}',
            'data': None
        }), 500


@trading_bp.route('/execution/plans/import', methods=['POST'])
def import_execution_plan():
    """
    导入执行方案接口
    
    请求体:
        {
            "plan_data": {
                "id": "plan_id",
                "name": "方案名称",
                "description": "方案描述",
                "combinations": [
                    {
                        "id": "combo_id",
                        "selection_strategy": "选股策略名称",
                        "timing_strategy": "择时策略名称",
                        "enabled": true
                    }
                ],
                "config_ref": "default",
                "created_at": "2024-01-01T12:00:00",
                "updated_at": "2024-01-01T12:00:00"
            }
        }
    
    返回:
        {
            "success": true/false,
            "message": "成功或错误信息",
            "data": {
                "plan_id": "plan_id"
            }
        }
    """
    try:
        from trading.strategy_execution_plan import ExecutionPlan
        
        data = request.get_json() or {}
        plan_data = data.get('plan_data')
        
        if not plan_data:
            return jsonify({
                'success': False,
                'message': '缺少方案数据',
                'data': None
            }), 400
        
        # 创建新方案，生成新的ID
        plan = ExecutionPlan.from_dict(plan_data)
        
        # 生成新ID避免冲突
        from uuid import uuid4
        plan.id = str(uuid4())
        plan.created_at = datetime.datetime.now().isoformat()
        plan.updated_at = datetime.datetime.now().isoformat()
        
        # 验证并保存
        if not plan.validate():
            return jsonify({
                'success': False,
                'message': '方案验证失败',
                'data': None
            }), 400
        
        plan.save()
        
        return jsonify({
            'success': True,
            'message': '导入执行方案成功',
            'data': {
                'plan_id': plan.id
            }
        }), 201
    
    except ValueError as e:
        return jsonify({
            'success': False,
            'message': str(e),
            'data': None
        }), 400
    except Exception as e:
        logger.error(f"导入执行方案失败: {str(e)}")
        return jsonify({
            'success': False,
            'message': f'导入执行方案失败: {str(e)}',
            'data': None
        }), 500


@trading_bp.route('/execution/plans/<plan_id>/export', methods=['GET'])
def export_execution_plan(plan_id):
    """
    导出执行方案接口
    
    参数:
        plan_id: 方案ID (路径参数)
    
    返回:
        JSON文件下载
    """
    try:
        from trading.strategy_execution_plan import ExecutionPlan
        from flask import make_response
        from urllib.parse import quote
        
        plan = ExecutionPlan.load(plan_id)
        
        if not plan:
            return jsonify({
                'success': False,
                'message': '方案不存在',
                'data': None
            }), 404
        
        # 导出为JSON
        plan_dict = plan.to_dict()
        
        response = make_response(json.dumps(plan_dict, ensure_ascii=False, indent=2))
        response.headers['Content-Type'] = 'application/json'
        response.headers['Content-Disposition'] = f'attachment; filename="{quote(f"{plan.name}.json")}"'
        
        return response
    
    except FileNotFoundError:
        return jsonify({
            'success': False,
            'message': '方案不存在',
            'data': None
        }), 404
    except Exception as e:
        logger.error(f"导出执行方案失败: {str(e)}")
        return jsonify({
            'success': False,
            'message': f'导出执行方案失败: {str(e)}',
            'data': None
        }), 500
